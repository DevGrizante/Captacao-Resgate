"""
Conciliação CVM x Quantum: que fundos a CVM lista e o relatório não cobre.

O QUE ESTE SCRIPT RESPONDE

    "O relatório diário do Quantum é a nossa visão do mercado. Ele está
    completo?"

    A resposta sai comparando duas listas de CNPJ:

        CVM     registro_fundo_classe.zip, fundos EM FUNCIONAMENTO NORMAL
        Quantum a coluna CNPJ da planilha vinculado_*.xlsx mais recente

    O que está na primeira e não na segunda é ponto cego: fundo vivo, com PL
    declarado, que a mesa não enxerga no painel.

POR QUE O PADRÃO NÃO É "TODOS OS FUNDOS DA CVM"

    A CVM lista ~34 mil fundos ativos, e a esmagadora maioria é ação,
    multimercado macro, cambial, previdência — nada que uma mesa de crédito
    privado negocie. Uma planilha com 30 mil linhas irrelevantes não é um
    achado, é ruído.

    Por isso o padrão é `--modo credito`: aplica os MESMOS sinais de crédito
    privado que o pipeline já usa para montar o universo do painel
    (services/credito_privado.py), sobre as MESMAS bases da CVM. O que sai é
    comparável com o que a tela mostra.

    `--modo todos` existe para auditoria — quando a pergunta for mesmo "o que
    mais existe lá fora", sem recorte.

SAÍDA

    data/relatorios/cnpjs_cvm_sem_quantum_<AAAAMMDD>.xlsx

    Aba "Faltantes"   uma linha por CNPJ ausente, com nome, gestor,
                      administrador, PL, classificação ANBIMA e quais sinais
                      de crédito dispararam — ordenada por PL desc, porque é
                      por aí que se decide o que perseguir primeiro.
    Aba "Resumo"      os números da conciliação (quantos de cada lado, quanto
                      de PL está fora) e os parâmetros da rodada, para a
                      planilha se explicar sozinha meses depois.
    Aba "Por gestora" o PL ausente somado por gestor, que é como a cobertura
                      costuma ser negociada — casa a casa, não fundo a fundo.

USO

    python backend/scripts/analise/conciliar_cnpjs.py
    python backend/scripts/analise/conciliar_cnpjs.py --modo todos
    python backend/scripts/analise/conciliar_cnpjs.py --quantum data/inbox/vinculado_20260820_0727.xlsx
    python backend/scripts/analise/conciliar_cnpjs.py --saida C:\\tmp\\conciliacao.xlsx
    python backend/scripts/analise/conciliar_cnpjs.py --pl-minimo 50000000

DESEMPENHO

    Tudo o que ele lê já é cache do pipeline (parquet em data/cache). Numa
    máquina com o cache quente a rodada leva segundos e NÃO toca a rede. Com
    cache frio ele baixa as mesmas bases que o painel baixaria — e o painel
    passa a aproveitar o download.
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.config import DATA_DIR, INBOX_DIR  # noqa: E402
from app.connectors import cvm_cadastro, cvm_documentos  # noqa: E402
from app.services import credito_privado  # noqa: E402
from app.utils import so_digitos  # noqa: E402

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s"
)
logger = logging.getLogger("conciliar")

SAIDA_DIR = DATA_DIR / "relatorios"

# Situação ativa no registro da CVM. Igual ao que cvm_cadastro usa; repetida
# aqui como constante nomeada porque este script filtra por ela diretamente.
ATIVO = cvm_cadastro.ATIVO


# =============================================================================
#  Lado Quantum
# =============================================================================

def _quantum_cnpjs(caminho: Path) -> set[str]:
    """CNPJs de fundo presentes na planilha do Quantum.

    Lê o arquivo bruto e localiza a coluna pelo NOME do cabeçalho, e não por
    posição: as duas colunas de CNPJ foram acrescentadas no meio da planilha em
    14/08/2026, e qualquer índice fixo passaria a ler a coluna errada em
    silêncio na próxima mudança de layout.

    Só a coluna "CNPJ" (do fundo) entra — "CNPJ Gestão" é da gestora e
    compará-la com o registro de fundos não faria sentido nenhum.
    """
    raw = pd.read_excel(caminho, sheet_name=0, header=None)

    coluna = None
    for col in range(raw.shape[1]):
        if _chave(raw.iat[0, col]) == "cnpj":
            coluna = col
            break

    if coluna is None:
        raise SystemExit(
            f"{caminho.name}: não achei a coluna 'CNPJ' no cabeçalho. Arquivos "
            "anteriores a 14/08/2026 não têm CNPJ e não servem para conciliar."
        )

    # Linha 4 em diante são dados (1 cabeçalho, 2 e 3 são as faixas de janela).
    valores = raw.iloc[3:, coluna].map(so_digitos)
    return {c for c in valores if c}


def _chave(valor) -> str:
    """'CNPJ Gestão' -> 'cnpj gestao'. Tolera acento, caixa e espaço extra."""
    s = unicodedata.normalize("NFKD", str(valor))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _planilha_mais_recente() -> Path:
    arquivos = sorted(INBOX_DIR.glob("vinculado_*.xlsx"))
    if not arquivos:
        raise SystemExit(
            f"Nenhuma planilha em {INBOX_DIR}. Rode o coletor, ou passe "
            "--quantum com o caminho do arquivo."
        )
    return arquivos[-1]


# =============================================================================
#  Lado CVM
# =============================================================================

def _cvm_universo(modo: str, limiar: float) -> pd.DataFrame:
    """Fundos ativos na CVM, já com os sinais de crédito privado avaliados.

    Devolve uma linha por CNPJ de fundo. As colunas de sinal vêm de duas bases
    extras (PERFIL_MENSAL e EXTRATO) porque são elas que carregam o sinal
    QUANTITATIVO — sem elas sobrariam só nome e classificação ANBIMA, e a
    conciliação herdaria a cegueira de quem só olha o nome do fundo.
    """
    cad = cvm_cadastro.carregar(apenas_ativos=True)
    cad = cad[cad["situacao"] == ATIVO].copy()
    logger.info("CVM: %d fundos em funcionamento normal.", len(cad))

    # As duas bases são indexadas pelo CNPJ da CLASSE, não do fundo. A ponte
    # existe no cadastro (`cnpj_classe`) e precisa ser feita antes do join,
    # senão a cobertura despenca sem motivo aparente.
    perfil = _seguro(cvm_documentos.perfil_mensal, "PERFIL_MENSAL")
    extrato = _seguro(cvm_documentos.extrato, "EXTRATO")

    cad["pct_credito_privado"] = _por_classe(cad, perfil, "pct_credito_privado")
    # `.eq(True)` NÃO é cosmético. O EXTRATO cobre ~24 mil das ~33 mil
    # classes; para o resto o `.map` devolve NaN — e `bool(float('nan'))` é
    # True em Python. Sem esta linha o sinal do extrato dispara para todo fundo
    # que a base não cobre, e a conciliação devolve o registro inteiro da CVM
    # como se fosse tudo crédito privado.
    cad["credito_privado"] = _por_classe(cad, extrato, "credito_privado").eq(True)

    marcados = []
    sinais = []
    for _cnpj, linha in cad.iterrows():
        pct = linha["pct_credito_privado"]
        fundo = {
            "nome": linha["nome"],
            "classificacao_anbima": linha["classificacao_anbima"],
            "pct_credito_privado": None if pd.isna(pct) else float(pct),
            "credito_privado": bool(linha["credito_privado"]),
        }
        ok, quais = credito_privado.avaliar(fundo, modo, limiar)
        marcados.append(ok)
        sinais.append(",".join(quais))

    cad["e_credito"] = marcados
    cad["sinais_credito"] = sinais
    return cad


def _seguro(fn, nome: str) -> pd.DataFrame:
    """Chama um carregador da CVM; base indisponível vira DataFrame vazio.

    Uma base fora do ar degrada a conciliação (menos sinais), mas não pode
    derrubá-la: o resultado com três sinais em vez de quatro ainda é útil, e
    quem roda precisa saber disso pelo log em vez de por um traceback.
    """
    try:
        df = fn()
        logger.info("%s: %d linhas.", nome, len(df))
        return df
    except Exception as e:  # noqa: BLE001
        logger.warning("%s indisponível (%s) — seguindo sem este sinal.", nome, e)
        return pd.DataFrame()


def _por_classe(cad: pd.DataFrame, base: pd.DataFrame, coluna: str):
    """Traz uma coluna indexada por CNPJ de classe para o índice de fundo."""
    if base.empty or coluna not in base.columns:
        return pd.Series(pd.NA, index=cad.index)
    return cad["cnpj_classe"].map(base[coluna])


# =============================================================================
#  Conciliação
# =============================================================================

def conciliar(
    caminho_quantum: Path,
    modo: str,
    limiar: float,
    pl_minimo: float,
) -> tuple[pd.DataFrame, dict]:
    """Devolve (faltantes, resumo)."""
    no_quantum = _quantum_cnpjs(caminho_quantum)
    logger.info("Quantum: %d CNPJs distintos em %s.", len(no_quantum), caminho_quantum.name)

    cvm = _cvm_universo(modo, limiar)
    universo = cvm[cvm["e_credito"]] if modo != "todos" else cvm
    logger.info("CVM no recorte '%s': %d fundos.", modo, len(universo))

    faltantes = universo[~universo.index.isin(no_quantum)].copy()

    if pl_minimo > 0:
        antes = len(faltantes)
        # `fillna(0)` de propósito: fundo sem PL declarado não passa no piso.
        # Ele existe e pode importar, mas não dá para priorizá-lo por tamanho,
        # e este corte é justamente uma ferramenta de priorização.
        faltantes = faltantes[faltantes["pl"].fillna(0) >= pl_minimo]
        logger.info(
            "Piso de PL de R$ %.0f: %d -> %d fundos.", pl_minimo, antes, len(faltantes)
        )

    faltantes = faltantes.sort_values("pl", ascending=False, na_position="last")

    # Interseção e "só no Quantum" não vão para a planilha, mas vão para o
    # resumo: sem eles não dá para saber se um número alto de faltantes
    # significa relatório incompleto ou recorte mal escolhido.
    intersecao = len(set(universo.index) & no_quantum)
    so_quantum = len(no_quantum - set(cvm.index))

    resumo = {
        "Gerado em": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "Planilha Quantum": caminho_quantum.name,
        "Modo do recorte": modo,
        "Limiar crédito privado (% do PL)": limiar,
        "Piso de PL (R$)": pl_minimo,
        "CNPJs no relatório Quantum": len(no_quantum),
        "Fundos ativos na CVM": len(cvm),
        f"Fundos da CVM no recorte '{modo}'": len(universo),
        "Presentes nos dois": intersecao,
        "FALTANDO no Quantum": len(faltantes),
        "No Quantum e fora do registro ativo da CVM": so_quantum,
        "PL total ausente (R$)": float(faltantes["pl"].fillna(0).sum()),
        "Faltantes sem PL declarado": int(faltantes["pl"].isna().sum()),
    }
    return faltantes, resumo


# =============================================================================
#  Saída
# =============================================================================

COLUNAS = {
    "cnpj": "CNPJ",
    "nome": "Nome do fundo",
    "gestor": "Gestor",
    "administrador": "Administrador",
    "pl": "PL (R$)",
    "pl_data": "Data do PL",
    "classificacao_anbima": "Classificação ANBIMA",
    "classificacao": "Classificação CVM",
    "forma_condominio": "Condomínio",
    "pct_credito_privado": "% em crédito privado (CVM)",
    "sinais_credito": "Sinais de crédito",
    "cnpj_classe": "CNPJ da classe",
}


def gravar(faltantes: pd.DataFrame, resumo: dict, destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)

    saida = faltantes.reset_index()[list(COLUNAS)].rename(columns=COLUNAS)

    por_gestora = (
        faltantes.assign(gestor=faltantes["gestor"].fillna("(sem gestor no registro)"))
        .groupby("gestor")
        .agg(fundos=("nome", "size"), pl=("pl", "sum"))
        .sort_values("pl", ascending=False)
        .reset_index()
        .rename(columns={"gestor": "Gestor", "fundos": "Fundos ausentes",
                         "pl": "PL ausente (R$)"})
    )

    with pd.ExcelWriter(destino, engine="openpyxl") as xls:
        saida.to_excel(xls, sheet_name="Faltantes", index=False)
        pd.DataFrame(
            {"Item": list(resumo), "Valor": [resumo[k] for k in resumo]}
        ).to_excel(xls, sheet_name="Resumo", index=False)
        por_gestora.to_excel(xls, sheet_name="Por gestora", index=False)

        _ajustar(xls.book["Faltantes"], saida)
        _ajustar(xls.book["Resumo"], None)
        _ajustar(xls.book["Por gestora"], por_gestora)

    logger.info("Gravado: %s (%d linhas).", destino, len(saida))


def _ajustar(aba, df: pd.DataFrame | None) -> None:
    """Largura de coluna e formato de moeda — a planilha é para ler, não para
    reprocessar; sair com '###' em toda coluna de PL derruba o valor dela."""
    from openpyxl.utils import get_column_letter

    aba.freeze_panes = "A2"
    for i, col in enumerate(aba.iter_cols(), start=1):
        largura = max(
            (len(str(c.value)) for c in col[:200] if c.value is not None), default=10
        )
        aba.column_dimensions[get_column_letter(i)].width = min(max(largura + 2, 10), 55)

    if df is None:
        return
    for j, nome in enumerate(df.columns, start=1):
        if "R$" not in nome:
            continue
        for linha in range(2, aba.max_row + 1):
            aba.cell(row=linha, column=j).number_format = "#,##0"


# =============================================================================
#  CLI
# =============================================================================

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--quantum", type=Path, help="planilha vinculado_*.xlsx a usar")
    p.add_argument("--saida", type=Path, help="caminho do .xlsx de saída")
    p.add_argument(
        "--modo", default="credito", choices=("credito", "estrito", "todos"),
        help="credito = qualquer sinal (padrão) | estrito = só o %% declarado à "
             "CVM | todos = sem recorte",
    )
    p.add_argument("--limiar", type=float, default=5.0,
                   help="%% mínimo do PL em crédito privado (padrão: 5)")
    p.add_argument("--pl-minimo", type=float, default=0.0,
                   help="descarta faltantes abaixo deste PL (R$)")
    args = p.parse_args()

    caminho = args.quantum or _planilha_mais_recente()
    if not caminho.exists():
        raise SystemExit(f"Planilha não encontrada: {caminho}")

    # "credito" é o nome amigável do modo que credito_privado chama de "sinal".
    modo = {"credito": "sinal"}.get(args.modo, args.modo)

    faltantes, resumo = conciliar(caminho, modo, args.limiar, args.pl_minimo)

    destino = args.saida or (
        SAIDA_DIR / f"cnpjs_cvm_sem_quantum_{datetime.now():%Y%m%d}.xlsx"
    )
    gravar(faltantes, resumo, destino)

    print()
    for chave, valor in resumo.items():
        if isinstance(valor, float) and valor > 1000:
            print(f"  {chave:<50} {valor:>18,.0f}".replace(",", "."))
        else:
            print(f"  {chave:<50} {valor:>18}")
    print(f"\n  -> {destino}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
