"""
Nível de negociação das DEBÊNTURES que os fundos carregam: CVM (CDA) x B3.

    python backend/scripts/analise/b3_liquidez.py                # 10 pregões
    python backend/scripts/analise/b3_liquidez.py --dias 20
    python backend/scripts/analise/b3_liquidez.py --saida C:\\tmp\\liquidez.xlsx

>>> SÓ DEBÊNTURE. LETRA FINANCEIRA FOI TIRADA DAQUI (20/08/2026)

    A LF não casa com a B3 e não vale a tentativa. O BLC_5 do CDA não traz
    `CD_ATIVO` nem `CD_ISIN` — a LF é identificada só por emissor, vencimento e
    taxa —, e o cadastro público da B3 só publica LF de oferta PÚBLICA sem
    esforço restrito, enquanto a LF que os fundos carregam é majoritariamente
    colocação privada. O casamento composto medido deu 1,0%.

    Se um dia isso voltar à mesa, o caminho que sobrava era o nível do EMISSOR
    (18,2% dos emissores, 61,7% do R$), nunca o do papel.

>>> A DEBÊNTURE, AO CONTRÁRIO, CASA DIRETO

    O BLC_4 traz `CD_ATIVO` e `CD_ISIN` preenchidos em 100% das linhas, e o
    `CD_ATIVO` É o código da B3, na grafia da B3. Medido no CDA de 2026-04
    contra o cadastro B3 de 2026-08-19: 2.005 de 2.145 papéis (93,5%) e 96,7%
    do R$ em carteira.

    O mesmo `CD_ATIVO` já é a chave que o projeto usa para buscar o indexador
    no SND (99,7% das posições). Ou seja, ele é uma chave TRIPLA:
    CVM (o que o fundo tem) x SND (indexador) x B3 (preço e liquidez).

>>> O EMISSOR DA DEBÊNTURE SÓ EXISTE DO LADO DA B3

    Detalhe que decide a arquitetura de qualquer tela por emissor: o BLC_4 NÃO
    tem `CNPJ_EMISSOR` nem `EMISSOR` (o `DS_ATIVO` vem como "DEB" ou
    "DEBENTURES SIMPLES", inútil). Quem carrega o nome da companhia emissora é
    o cadastro da B3, alcançado pelo ticker.

    Sem B3, o melhor disponível é o prefixo de 4 letras do ticker (SBSP13 ->
    SBSP), que agrupa 2.160 tickers em 886 emissores — serve para agregar, mas
    não dá nome nem CNPJ.

>>> O QUE É "NÍVEL DE NEGOCIAÇÃO" AQUI

    Não é volume. Volume alto num único dia costuma ser transferência, não
    liquidez. As medidas que descrevem se dá para montar ou desmontar posição:

        dias_negociados / pregões    com que FREQUÊNCIA o papel negocia
        negocios_extragrupo          negócio entre partes independentes
        pct_extragrupo               quanto do fluxo é mercado de verdade
        amplitude_pct                (máximo − mínimo) / médio, proxy de spread

    A distinção INTRAGRUPO x EXTRAGRUPO é a mais importante e a mais ignorada.
    Num mês de pregões, das 29.137 linhas de debênture negociada, 19.503 eram
    INTRAGRUPO. Um painel que somasse volume sem separar isso mostraria uma
    liquidez que não existe.

>>> SAÍDA

    data/relatorios/b3_liquidez_<data>.xlsx

    Aba "Debentures"  papel a papel, com o R$ que os fundos carregam ao lado do
                      nível de negociação. Ordenada pelo R$ em carteira: o que
                      importa primeiro é o que temos, não o que gira.
    Aba "Cobertura"   quanto casou e com que régua — para a planilha se
                      explicar sozinha meses depois.
"""
from __future__ import annotations

import argparse
import logging
import sys
import zipfile
from datetime import date, timedelta
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import b3_balcao  # noqa: E402
from app.config import DATA_DIR  # noqa: E402
from app.connectors import cvm_cda_arquivo  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
logger = logging.getLogger("b3_liquidez")

SAIDA_DIR = DATA_DIR / "relatorios"


def _num(serie) -> pd.Series:
    """'1.014,55' -> 1014.55. A B3 exporta no formato brasileiro."""
    return pd.to_numeric(
        serie.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
        errors="coerce",
    )


def _coluna(df: pd.DataFrame, trecho: str) -> str:
    """Acha a coluna pelo trecho do nome — os cabeçalhos da B3 vêm acentuados."""
    for c in df.columns:
        if trecho.lower() in c.lower():
            return c
    raise KeyError(f"coluna com {trecho!r} não encontrada em {list(df.columns)}")


# =============================================================================
#  Lado B3
# =============================================================================

def _pregoes(dias: int) -> list[str]:
    """Os N últimos dias úteis, do mais recente para trás."""
    d, out = date.today(), []
    while len(out) < dias:
        d -= timedelta(days=1)
        if d.weekday() < 5:
            out.append(d.isoformat())
    return out


def negociacao(dias: int) -> pd.DataFrame:
    """Negociação consolidada acumulada em N pregões, papel a papel.

    Um dia só não descreve liquidez: um papel que negocia toda semana pode não
    ter negociado hoje, e apareceria como ilíquido. A janela é o que separa
    "não negocia" de "não negociou hoje".
    """
    partes = []
    for dia in _pregoes(dias):
        try:
            df = b3_balcao.exportar("ConsolidatedRecords", dia)
        except Exception as e:  # noqa: BLE001 — um pregão faltando não invalida a janela
            logger.warning("Pregão %s indisponível (%s).", dia, e)
            continue
        if df.empty:
            continue
        # Filtra para DEB o quanto antes: o arquivo do dia tem ~73 mil linhas e
        # só ~2,5% são debênture. Concatenar tudo para descartar depois custaria
        # memória por nada.
        c_tipo = _coluna(df, "Instrumento financeiro")
        df = df[df[c_tipo] == "DEB"].copy()
        if df.empty:
            continue
        df["_dia"] = dia
        partes.append(df)
        logger.info("Pregão %s: %d linhas de debênture.", dia, len(df))

    if not partes:
        return pd.DataFrame()

    con = pd.concat(partes, ignore_index=True)
    con["ticker"] = con[_coluna(con, "Código IF")].astype(str).str.strip().str.upper()
    con["emissor_b3"] = con[_coluna(con, "Emissor")].astype(str).str.strip()
    con["volume"] = _num(con[_coluna(con, "Volume financeiro")])
    con["negocios"] = _num(con[_coluna(con, "Número de negócios")])
    con["p_min"] = _num(con[_coluna(con, "Preço mínimo")])
    con["p_max"] = _num(con[_coluna(con, "Preço máximo")])
    con["p_med"] = _num(con[_coluna(con, "Preço médio")])
    # INTRAGRUPO é transferência entre partes ligadas. Contar isso como
    # liquidez é o erro que faz um papel travado parecer negociável.
    con["extragrupo"] = (
        con[_coluna(con, "Classificação")].astype(str).str.upper().str.startswith("EXTRA")
    )

    g = con.groupby("ticker")
    out = pd.DataFrame({
        "emissor_b3": g["emissor_b3"].first(),
        "dias_negociados": g["_dia"].nunique(),
        "negocios": g["negocios"].sum(),
        "volume": g["volume"].sum(),
        "preco_medio": g["p_med"].mean(),
        "p_min": g["p_min"].min(),
        "p_max": g["p_max"].max(),
    })
    extra = con[con["extragrupo"]].groupby("ticker")
    out["dias_extragrupo"] = extra["_dia"].nunique()
    out["negocios_extragrupo"] = extra["negocios"].sum()
    out["volume_extragrupo"] = extra["volume"].sum()
    out[["dias_extragrupo", "negocios_extragrupo", "volume_extragrupo"]] = (
        out[["dias_extragrupo", "negocios_extragrupo", "volume_extragrupo"]].fillna(0)
    )

    out["pregoes"] = con["_dia"].nunique()
    out["freq_negociacao"] = out["dias_extragrupo"] / out["pregoes"]
    out["pct_extragrupo"] = (out["volume_extragrupo"] / out["volume"]).where(out["volume"] > 0)
    # Amplitude no período como proxy de spread. Não é bid-ask (a B3 não
    # publica o book do balcão), mas descreve o mesmo: quanto o preço anda
    # entre dois negócios do mesmo papel.
    out["amplitude_pct"] = ((out["p_max"] - out["p_min"]) / out["preco_medio"] * 100).where(
        out["preco_medio"] > 0
    )
    return out


def nivel(linha) -> str:
    """Rótulo de liquidez. A frequência EXTRAGRUPO manda — o resto é detalhe.

    Os cortes são de negócio, não estatísticos: negociar em mais de metade dos
    pregões é o que a mesa chama de papel líquido; negociar em nenhum é papel
    que só sai no vencimento, por mais volume intragrupo que tenha girado.
    """
    freq = linha.get("freq_negociacao") or 0
    if freq >= 0.5:
        return "Líquido"
    if freq >= 0.2:
        return "Moderado"
    if freq > 0:
        return "Esporádico"
    if (linha.get("dias_negociados") or 0) > 0:
        return "Só intragrupo"
    return "Sem negócio"


# =============================================================================
#  Lado CVM
# =============================================================================

def debentures_em_carteira() -> tuple[pd.DataFrame, str]:
    """As debêntures que os fundos carregam, por ticker. (tabela, mês do CDA)."""
    mes = cvm_cda_arquivo.mes_alvo()
    z = cvm_cda_arquivo.abrir(mes)
    if z is None:
        raise SystemExit(f"CDA {mes} indisponível.")
    with z:
        return _ler_blc4(z, mes), mes


def _ler_blc4(z: zipfile.ZipFile, mes: str) -> pd.DataFrame:
    df = pd.read_csv(
        z.open(f"cda_fi_BLC_4_{mes}.csv"), sep=";", encoding="latin-1", low_memory=False,
        usecols=["CNPJ_FUNDO_CLASSE", "TP_ATIVO", "CD_ATIVO", "CD_ISIN",
                 "VL_MERC_POS_FINAL", "DT_FIM_VIGENCIA"],
    )
    df = df[df["TP_ATIVO"].astype(str).str.contains("nture", na=False)].copy()
    df["ticker"] = df["CD_ATIVO"].astype(str).str.strip().str.upper()
    df["valor"] = pd.to_numeric(df["VL_MERC_POS_FINAL"], errors="coerce")
    df = df[df["valor"] > 0]

    g = df.groupby("ticker")
    out = pd.DataFrame({
        "isin": g["CD_ISIN"].first(),
        "vencimento": pd.to_datetime(g["DT_FIM_VIGENCIA"].first(), errors="coerce"),
        "fundos": g["CNPJ_FUNDO_CLASSE"].nunique(),
        "valor_em_carteira": g["valor"].sum(),
    })
    # Prefixo de 4 letras = código do emissor no mercado de debêntures
    # (SBSP13 -> SBSP). É o agrupador que existe SEM a B3; o nome da companhia
    # só vem de lá, porque o BLC_4 não traz emissor nenhum.
    out["prefixo"] = out.index.str[:4]
    return out.sort_values("valor_em_carteira", ascending=False)


# =============================================================================
#  Junção
# =============================================================================

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--dias", type=int, default=10, help="pregões na janela (padrão: 10)")
    p.add_argument("--saida", type=Path)
    args = p.parse_args()

    logger.info("Lendo o CDA da CVM…")
    deb, mes = debentures_em_carteira()
    logger.info("CDA %s: %d debêntures distintas em carteira.", mes, len(deb))

    logger.info("Baixando o cadastro de instrumentos da B3…")
    dia_cad = b3_balcao._data_boa()
    cad = b3_balcao.exportar("InstrumentRegistration", dia_cad)
    cad["ticker"] = cad[_coluna(cad, "Código IF")].astype(str).str.strip().str.upper()
    b3_deb = cad[cad[_coluna(cad, "Instrumento financeiro")] == "DEB"]
    b3_deb = b3_deb.drop_duplicates("ticker").set_index("ticker")

    logger.info("Baixando %d pregões de negociação…", args.dias)
    neg = negociacao(args.dias)
    logger.info("Negociação: %d debêntures com negócio na janela.", len(neg))

    renomear = {
        _coluna(cad, "Emissor"): "emissor",
        _coluna(cad, "Indexador"): "indexador",
        _coluna(cad, "Taxa adicional"): "taxa",
        _coluna(cad, "Incentivada"): "incentivada",
    }
    d = deb.join(b3_deb[list(renomear)], how="left").rename(columns=renomear)
    d["no_cadastro_b3"] = d.index.isin(b3_deb.index)
    d = d.join(neg.drop(columns=["emissor_b3"]), how="left")
    d["dias_negociados"] = d["dias_negociados"].fillna(0)
    d["freq_negociacao"] = d["freq_negociacao"].fillna(0)
    d["nivel"] = d.apply(nivel, axis=1)

    cobertura = _cobertura(d, mes, dia_cad, neg)
    destino = args.saida or (SAIDA_DIR / f"b3_liquidez_{date.today():%Y%m%d}.xlsx")
    _gravar(d, cobertura, destino)

    print()
    for k, v in cobertura.items():
        print(f"  {k:<52} {v}")
    print(f"\n  -> {destino}")
    return 0


def _cobertura(d, mes, dia_cad, neg) -> dict:
    val = d["valor_em_carteira"]
    casou = d["no_cadastro_b3"]
    negociou = d["freq_negociacao"] > 0
    return {
        "CDA de referência": mes,
        "Cadastro B3 de": dia_cad,
        "Pregões na janela": int(neg["pregoes"].iloc[0]) if not neg.empty else 0,
        "Debêntures em carteira": len(d),
        "  casadas no cadastro B3": f"{int(casou.sum())} ({casou.mean():.1%})",
        "  R$ casado": f"{val[casou].sum() / 1e9:.1f} bi de {val.sum() / 1e9:.1f} bi "
                      f"({val[casou].sum() / val.sum():.1%})",
        "  com negócio EXTRAGRUPO na janela": f"{int(negociou.sum())} ({negociou.mean():.1%})",
        "  R$ em papel que negocia": f"{val[negociou].sum() / 1e9:.1f} bi "
                                     f"({val[negociou].sum() / val.sum():.1%})",
        "Emissores (prefixo do ticker)": d["prefixo"].nunique(),
    }


COLS = {
    "ticker": "Ticker (B3)", "isin": "ISIN", "emissor": "Emissor", "prefixo": "Cód. emissor",
    "indexador": "Indexador", "taxa": "Taxa", "incentivada": "Incentivada",
    "vencimento": "Vencimento", "fundos": "Fundos que carregam",
    "valor_em_carteira": "Em carteira (R$)", "no_cadastro_b3": "No cadastro B3",
    "nivel": "Nível de negociação", "freq_negociacao": "Freq. (dias c/ negócio)",
    "dias_negociados": "Dias com negócio", "negocios_extragrupo": "Negócios extragrupo",
    "volume_extragrupo": "Volume extragrupo (R$)", "pct_extragrupo": "% extragrupo",
    "amplitude_pct": "Amplitude %", "preco_medio": "Preço médio",
}


def _gravar(d, cobertura, destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    plano = d.reset_index()
    saida = plano[[c for c in COLS if c in plano.columns]].rename(columns=COLS)
    resumo = pd.DataFrame({"Item": list(cobertura), "Valor": [str(v) for v in cobertura.values()]})

    with pd.ExcelWriter(destino, engine="openpyxl") as xls:
        saida.to_excel(xls, sheet_name="Debentures", index=False)
        resumo.to_excel(xls, sheet_name="Cobertura", index=False)
        for aba in ("Debentures", "Cobertura"):
            _ajustar(xls.book[aba])
    logger.info("Gravado: %s (%d debêntures).", destino, len(saida))


def _ajustar(aba) -> None:
    from openpyxl.utils import get_column_letter

    aba.freeze_panes = "A2"
    for i, col in enumerate(aba.iter_cols(), start=1):
        largura = max((len(str(c.value)) for c in col[:200] if c.value is not None), default=10)
        aba.column_dimensions[get_column_letter(i)].width = min(max(largura + 2, 12), 46)
    cabecalho = [c.value for c in next(aba.iter_rows(max_row=1))]
    for j, nome in enumerate(cabecalho, start=1):
        if "R$" in str(nome):
            formato = "#,##0"
        elif "%" in str(nome) or "Freq" in str(nome):
            formato = "0.0%"
        else:
            continue
        for linha in range(2, aba.max_row + 1):
            aba.cell(row=linha, column=j).number_format = formato


if __name__ == "__main__":
    raise SystemExit(main())
